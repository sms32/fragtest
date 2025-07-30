"""
Excel QA Validator - Excel Parser Service
Advanced Excel parsing service that handles complex structures, merged cells, and multi-section layouts
"""

import openpyxl
from openpyxl import load_workbook
from typing import Dict, List, Tuple, Any, Optional
import pandas as pd
import logging
import re
from dataclasses import dataclass
from app.models.comparison_models import ParsedExcelData, ValidationError

logger = logging.getLogger(__name__)

@dataclass
class CellInfo:
    """Information about a cell including merged cell handling"""
    value: Any
    row: int
    col: int
    is_merged: bool
    merged_range: Optional[str] = None

class ExcelStructureDetector:
    """Detects structure patterns in Excel files"""
    
    def __init__(self):
        self.section_keywords = {
            'BQ': ['baqala', 'bq', 'baQALA'],
            'NA': ['national accounts', 'na', 'national_accounts'],
            'COMBINED': ['total', 'combined', 'central', 'summary']
        }
        
        self.header_keywords = [
            'region', 'supervisor', 'area', 'wk slab', 'day sale', 'day slab', 
            'day stale', 'stale %', 'wtd slab', 'wtd sale', 'wtd ach%', 
            'wtd stale', 'stale%', 'wk sale ly', 'wk grw%', 'ytd ly', 'ytd ty', 'grw%'
        ]
    
    def detect_sections(self, worksheet) -> Dict[str, Tuple[int, int]]:
        """Detect section boundaries in the worksheet"""
        sections = {}
        current_section = None
        
        for row in range(1, worksheet.max_row + 1):
            for col in range(1, worksheet.max_column + 1):
                cell_value = self._get_cell_value_safe(worksheet, row, col)
                
                if not cell_value:
                    continue
                
                cell_text = str(cell_value).lower().strip()
                
                # Check for section keywords
                for section_type, keywords in self.section_keywords.items():
                    if any(keyword in cell_text for keyword in keywords):
                        # End previous section
                        if current_section:
                            sections[current_section] = (sections[current_section][0], row - 1)
                        
                        # Start new section
                        current_section = section_type
                        sections[section_type] = (row, None)
                        logger.debug(f"Found section '{section_type}' starting at row {row}")
                        break
        
        # Close the last section
        if current_section and sections[current_section][1] is None:
            sections[current_section] = (sections[current_section][0], worksheet.max_row)
        
        return sections
    
    def detect_headers(self, worksheet, start_row: int, end_row: int) -> Dict[int, str]:
        """Detect column headers in a section"""
        headers = {}
        header_candidates = {}
        
        # Look for headers in the first few rows of the section
        search_rows = min(5, end_row - start_row + 1)
        
        for row in range(start_row, start_row + search_rows):
            for col in range(1, worksheet.max_column + 1):
                cell_value = self._get_cell_value_safe(worksheet, row, col)
                
                if not cell_value:
                    continue
                
                cell_text = str(cell_value).lower().strip()
                
                # Check if this looks like a header
                for header_keyword in self.header_keywords:
                    if header_keyword in cell_text:
                        # Store the original case version
                        original_value = str(cell_value).strip()
                        if col not in header_candidates or len(original_value) > len(header_candidates[col]):
                            header_candidates[col] = original_value
        
        # Also look for numeric pattern headers (like percentages, numbers)
        for row in range(start_row, start_row + search_rows):
            for col in range(1, worksheet.max_column + 1):
                if col in header_candidates:
                    continue
                    
                cell_value = self._get_cell_value_safe(worksheet, row, col)
                if cell_value and str(cell_value).strip():
                    cell_text = str(cell_value).strip()
                    
                    # Look for patterns that suggest headers (not pure data)
                    if (not cell_text.replace(',', '').replace('.', '').replace('-', '').replace('(', '').replace(')', '').replace('%', '').isdigit() or
                        '%' in cell_text or
                        any(char.isalpha() for char in cell_text)):
                        header_candidates[col] = cell_text
        
        return header_candidates
    
    def _get_cell_value_safe(self, worksheet, row: int, col: int) -> Any:
        """Safely get cell value handling merged cells"""
        try:
            cell = worksheet.cell(row=row, column=col)
            
            # Check if cell is part of a merged range
            for merged_range in worksheet.merged_cells.ranges:
                if cell.coordinate in merged_range:
                    # Get value from top-left cell of merged range
                    return worksheet[merged_range.start_cell.coordinate].value
            
            return cell.value
        except Exception as e:
            logger.warning(f"Error reading cell at row {row}, col {col}: {e}")
            return None

class ExcelDataExtractor:
    """Extracts data from Excel worksheets"""
    
    def __init__(self):
        self.structure_detector = ExcelStructureDetector()
    
    def extract_data_rows(self, worksheet, start_row: int, end_row: int, headers: Dict[int, str]) -> List[Dict[str, Any]]:
        """Extract data rows from a section"""
        data_rows = []
        
        for row in range(start_row, end_row + 1):
            row_data = {}
            has_meaningful_data = False
            
            for col, header in headers.items():
                cell_value = self._get_processed_cell_value(worksheet, row, col)
                row_data[header] = cell_value
                
                # Check if this row has meaningful data
                if cell_value is not None and str(cell_value).strip():
                    has_meaningful_data = True
            
            # Only include rows that seem to contain actual data
            if has_meaningful_data and self._is_data_row(row_data):
                # Create composite key
                composite_key = self._create_composite_key(row_data)
                row_data['composite_key'] = composite_key
                
                # Clean up the data
                cleaned_data = self._clean_row_data(row_data)
                data_rows.append(cleaned_data)
        
        return data_rows
    
    def _get_processed_cell_value(self, worksheet, row: int, col: int) -> Any:
        """Get cell value with processing for different data types"""
        raw_value = self.structure_detector._get_cell_value_safe(worksheet, row, col)
        
        if raw_value is None:
            return None
        
        # Convert to string first
        str_value = str(raw_value).strip()
        
        if not str_value:
            return None
        
        # Try to convert numbers
        if self._is_numeric(str_value):
            return self._parse_numeric(str_value)
        
        # Handle percentages
        if str_value.endswith('%'):
            try:
                num_part = str_value[:-1].replace(',', '')
                return f"{float(num_part)}%"
            except:
                return str_value
        
        # Handle parentheses (negative numbers)
        if str_value.startswith('(') and str_value.endswith(')'):
            inner_value = str_value[1:-1]
            if self._is_numeric(inner_value):
                return f"({self._parse_numeric(inner_value)})"
            return str_value
        
        return str_value
    
    def _is_numeric(self, value: str) -> bool:
        """Check if a string represents a number"""
        # Remove commas and check if it's a number
        clean_value = value.replace(',', '').replace('-', '')
        try:
            float(clean_value)
            return True
        except ValueError:
            return False
    
    def _parse_numeric(self, value: str) -> float:
        """Parse numeric value from string"""
        try:
            # Remove commas and parse
            clean_value = value.replace(',', '')
            return float(clean_value)
        except ValueError:
            return value  # Return original if parsing fails
    
    def _is_data_row(self, row_data: Dict[str, Any]) -> bool:
        """Determine if a row contains actual data (not headers or empty rows)"""
        # Check for key identifying fields
        key_fields = ['Region', 'Supervisor', 'Area']
        
        # Look for variations of these field names
        region_values = []
        supervisor_values = []
        area_values = []
        
        for field, value in row_data.items():
            if not value:
                continue
                
            field_lower = field.lower()
            value_str = str(value).strip()
            
            if 'region' in field_lower and value_str:
                region_values.append(value_str)
            elif 'supervis' in field_lower and value_str:
                supervisor_values.append(value_str)
            elif 'area' in field_lower and value_str:
                area_values.append(value_str)
        
        # Check if we have meaningful identifiers
        has_identifiers = any([
            any('central' in str(v).lower() for v in region_values),
            any(str(v).lower() in ['michael', 'sms', 'mic', 's'] for v in supervisor_values),
            any(str(v).lower() in ['mgg', 'gmg', 'mg', 'm'] for v in area_values)
        ])
        
        # Also check for numeric data indicating this is a data row
        has_numeric_data = any(
            isinstance(value, (int, float)) for value in row_data.values()
        )
        
        return has_identifiers or has_numeric_data
    
    def _create_composite_key(self, row_data: Dict[str, Any]) -> str:
        """Create composite key from row data"""
        region = ""
        supervisor = ""
        area = ""
        
        for field, value in row_data.items():
            if not value:
                continue
                
            field_lower = field.lower()
            value_str = str(value).strip()
            
            if 'region' in field_lower:
                region = value_str
            elif 'supervis' in field_lower:
                supervisor = value_str
            elif 'area' in field_lower:
                area = value_str
        
        # Create composite key
        key_parts = [region, supervisor, area]
        composite_key = "_".join([part for part in key_parts if part]).replace(" ", "_")
        
        return composite_key if composite_key else f"row_{hash(str(row_data))}"
    
    def _clean_row_data(self, row_data: Dict[str, Any]) -> Dict[str, Any]:
        """Clean and standardize row data"""
        cleaned = {}
        
        for field, value in row_data.items():
            # Standardize field names
            clean_field = field.strip()
            
            # Clean values
            if value is None:
                cleaned[clean_field] = None
            elif isinstance(value, str):
                cleaned[clean_field] = value.strip() if value.strip() else None
            else:
                cleaned[clean_field] = value
        
        return cleaned

class ExcelParser:
    """Main Excel parser class"""
    
    def __init__(self):
        self.structure_detector = ExcelStructureDetector()
        self.data_extractor = ExcelDataExtractor()
    
    def parse_excel_file(self, file_path: str) -> ParsedExcelData:
        """Parse Excel file and return structured data"""
        logger.info(f"Parsing Excel file: {file_path}")
        
        try:
            # Load workbook
            workbook = load_workbook(file_path, data_only=True)
            worksheet = workbook.active
            
            logger.info(f"Worksheet loaded: {worksheet.max_row} rows, {worksheet.max_column} columns")
            
            # Detect sections
            sections = self.structure_detector.detect_sections(worksheet)
            logger.info(f"Detected sections: {list(sections.keys())}")
            
            parsed_sections = {}
            headers_info = {}
            parsing_errors = []
            total_records = 0
            
            # Process each section
            for section_name, (start_row, end_row) in sections.items():
                try:
                    logger.info(f"Processing section '{section_name}' (rows {start_row}-{end_row})")
                    
                    # Detect headers for this section
                    headers = self.structure_detector.detect_headers(worksheet, start_row, end_row)
                    headers_info[section_name] = headers
                    
                    if not headers:
                        logger.warning(f"No headers detected for section '{section_name}'")
                        parsing_errors.append(f"No headers detected for section '{section_name}'")
                        continue
                    
                    logger.debug(f"Headers for {section_name}: {headers}")
                    
                    # Extract data rows (skip a few rows after section start to avoid headers)
                    data_start_row = start_row + 3  # Skip section title and header rows
                    if data_start_row <= end_row:
                        data_rows = self.data_extractor.extract_data_rows(
                            worksheet, data_start_row, end_row, headers
                        )
                        
                        parsed_sections[section_name] = data_rows
                        total_records += len(data_rows)
                        
                        logger.info(f"Extracted {len(data_rows)} records from section '{section_name}'")
                    else:
                        logger.warning(f"No data rows found in section '{section_name}'")
                        parsed_sections[section_name] = []
                
                except Exception as e:
                    error_msg = f"Error processing section '{section_name}': {str(e)}"
                    logger.error(error_msg)
                    parsing_errors.append(error_msg)
                    parsed_sections[section_name] = []
            
            # Create metadata
            metadata = {
                'file_path': file_path,
                'worksheet_name': worksheet.title,
                'total_rows': worksheet.max_row,
                'total_columns': worksheet.max_column,
                'sections_found': list(sections.keys()),
                'parsing_timestamp': pd.Timestamp.now().isoformat()
            }
            
            logger.info(f"Parsing completed. Total records: {total_records}")
            
            return ParsedExcelData(
                sections=parsed_sections,
                headers=headers_info,
                metadata=metadata,
                total_records=total_records,
                parsing_errors=parsing_errors
            )
            
        except Exception as e:
            error_msg = f"Failed to parse Excel file {file_path}: {str(e)}"
            logger.error(error_msg)
            
            return ParsedExcelData(
                sections={},
                headers={},
                metadata={'file_path': file_path, 'error': error_msg},
                total_records=0,
                parsing_errors=[error_msg]
            )
    
    def validate_file_structure(self, file_path: str) -> List[ValidationError]:
        """Validate that the Excel file has the expected structure"""
        errors = []
        
        try:
            workbook = load_workbook(file_path, data_only=True)
            worksheet = workbook.active
            
            # Check if file is empty
            if worksheet.max_row <= 1:
                errors.append(ValidationError(
                    error_type="EMPTY_FILE",
                    message="Excel file appears to be empty or has only headers",
                    suggestion="Ensure the file contains data rows"
                ))
            
            # Check for sections
            sections = self.structure_detector.detect_sections(worksheet)
            if not sections:
                errors.append(ValidationError(
                    error_type="NO_SECTIONS",
                    message="No recognizable sections found (BQ, NA, etc.)",
                    suggestion="Ensure the file contains section headers like 'Baqala (BQ)' or 'National Accounts (NA)'"
                ))
            
            # Check for headers in each section
            for section_name, (start_row, end_row) in sections.items():
                headers = self.structure_detector.detect_headers(worksheet, start_row, end_row)
                if not headers:
                    errors.append(ValidationError(
                        error_type="NO_HEADERS",
                        message=f"No headers detected in section '{section_name}'",
                        field=section_name,
                        suggestion="Ensure the section contains column headers like 'Region', 'Supervisor', 'Area', etc."
                    ))
                elif len(headers) < 3:
                    errors.append(ValidationError(
                        error_type="INSUFFICIENT_HEADERS",
                        message=f"Section '{section_name}' has only {len(headers)} headers, expected at least 3",
                        field=section_name,
                        suggestion="Ensure the section has at least Region, Supervisor, and Area columns"
                    ))
            
        except Exception as e:
            errors.append(ValidationError(
                error_type="FILE_READ_ERROR",
                message=f"Cannot read Excel file: {str(e)}",
                suggestion="Check if the file is a valid Excel file and is not corrupted"
            ))
        
        return errors
    
    def get_file_preview(self, file_path: str, max_rows: int = 10) -> Dict[str, Any]:
        """Get a preview of the Excel file structure"""
        try:
            workbook = load_workbook(file_path, data_only=True)
            worksheet = workbook.active
            
            # Get basic info
            preview = {
                'filename': file_path.split('/')[-1],
                'worksheet_name': worksheet.title,
                'total_rows': worksheet.max_row,
                'total_columns': worksheet.max_column,
                'sections': {},
                'sample_data': []
            }
            
            # Detect sections
            sections = self.structure_detector.detect_sections(worksheet)
            
            for section_name, (start_row, end_row) in sections.items():
                headers = self.structure_detector.detect_headers(worksheet, start_row, end_row)
                preview['sections'][section_name] = {
                    'start_row': start_row,
                    'end_row': end_row,
                    'headers': headers,
                    'header_count': len(headers)
                }
            
            # Get sample data (first few rows)
            for row in range(1, min(max_rows + 1, worksheet.max_row + 1)):
                row_data = []
                for col in range(1, min(10, worksheet.max_column + 1)):  # First 10 columns
                    value = self.structure_detector._get_cell_value_safe(worksheet, row, col)
                    row_data.append(str(value) if value is not None else "")
                preview['sample_data'].append(row_data)
            
            return preview
            
        except Exception as e:
            return {
                'error': f"Cannot preview file: {str(e)}",
                'filename': file_path.split('/')[-1]
            }

# Utility functions
def is_excel_file(filename: str) -> bool:
    """Check if filename has Excel extension"""
    return filename.lower().endswith(('.xlsx', '.xls'))

def get_excel_file_info(file_path: str) -> Dict[str, Any]:
    """Get basic information about Excel file"""
    try:
        workbook = load_workbook(file_path, data_only=True)
        worksheet = workbook.active
        
        return {
            'valid': True,
            'worksheet_count': len(workbook.worksheets),
            'active_worksheet': worksheet.title,
            'max_row': worksheet.max_row,
            'max_column': worksheet.max_column,
            'has_merged_cells': len(worksheet.merged_cells.ranges) > 0,
            'merged_cell_count': len(worksheet.merged_cells.ranges)
        }
    except Exception as e:
        return {
            'valid': False,
            'error': str(e)
        }