"""
Excel QA Validator - File Utilities
Utility functions for file operations, storage management, and report handling
"""

import os
import shutil
import aiofiles
from typing import List, Dict, Any, Optional, Tuple
from datetime import datetime
import logging
from pathlib import Path
import pandas as pd
from openpyxl import Workbook
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side

from app.models.comparison_models import (
    ReportSummary, FileInfo, ComparisonResult, CalculationValidation,
    ValidationSummary, SectionSummary, get_file_info
)

logger = logging.getLogger(__name__)

class FileManager:
    """Manages file operations for reports"""
    
    def __init__(self, base_path: str = "./reports"):
        self.base_path = Path(base_path)
        self.base_path.mkdir(exist_ok=True)
        
        # Ensure base directory exists
        logger.info(f"FileManager initialized with base path: {self.base_path}")
    
    async def save_uploaded_file(self, file_content: bytes, report_name: str, file_type: str) -> str:
        """Save uploaded file to the reports directory"""
        
        # Create report directory
        report_dir = self.base_path / report_name
        report_dir.mkdir(exist_ok=True)
        
        # Determine filename
        filename = f"{file_type}.xlsx"
        file_path = report_dir / filename
        
        # Save file
        async with aiofiles.open(file_path, 'wb') as f:
            await f.write(file_content)
        
        logger.info(f"Saved {file_type} file for report '{report_name}': {file_path}")
        return str(file_path)
    
    def get_report_files(self, report_name: str) -> Tuple[Optional[str], Optional[str]]:
        """Get source and destination file paths for a report"""
        report_dir = self.base_path / report_name
        
        if not report_dir.exists():
            return None, None
        
        source_path = report_dir / "source.xlsx"
        dest_path = report_dir / "dest.xlsx"
        
        source_file = str(source_path) if source_path.exists() else None
        dest_file = str(dest_path) if dest_path.exists() else None
        
        return source_file, dest_file
    
    def list_available_reports(self) -> List[ReportSummary]:
        """List all available reports with their file information"""
        reports = []
        
        if not self.base_path.exists():
            return reports
        
        for report_dir in self.base_path.iterdir():
            if report_dir.is_dir():
                report_name = report_dir.name
                
                # Get file information
                source_path = report_dir / "source.xlsx"
                dest_path = report_dir / "dest.xlsx"
                
                source_info = get_file_info(str(source_path), "source.xlsx") if source_path.exists() else None
                dest_info = get_file_info(str(dest_path), "dest.xlsx") if dest_path.exists() else None
                
                # Get creation time (earliest file timestamp)
                created_at = datetime.now()
                if source_info and source_info.exists:
                    created_at = min(created_at, source_info.upload_timestamp)
                if dest_info and dest_info.exists:
                    created_at = min(created_at, dest_info.upload_timestamp)
                
                has_both_files = (source_info and source_info.exists) and (dest_info and dest_info.exists)
                
                reports.append(ReportSummary(
                    report_name=report_name,
                    source_file=source_info,
                    dest_file=dest_info,
                    created_at=created_at,
                    has_both_files=has_both_files
                ))
        
        # Sort by creation time (newest first)
        reports.sort(key=lambda x: x.created_at, reverse=True)
        
        return reports
    
    def delete_report(self, report_name: str) -> bool:
        """Delete a report and all its files"""
        report_dir = self.base_path / report_name
        
        if not report_dir.exists():
            return False
        
        try:
            shutil.rmtree(report_dir)
            logger.info(f"Deleted report: {report_name}")
            return True
        except Exception as e:
            logger.error(f"Failed to delete report {report_name}: {e}")
            return False
    
    def get_file_size_mb(self, file_path: str) -> float:
        """Get file size in MB"""
        try:
            size_bytes = os.path.getsize(file_path)
            return round(size_bytes / (1024 * 1024), 2)
        except:
            return 0.0
    
    def cleanup_old_reports(self, days_old: int = 30) -> int:
        """Clean up reports older than specified days"""
        if not self.base_path.exists():
            return 0
        
        cutoff_date = datetime.now().timestamp() - (days_old * 24 * 60 * 60)
        deleted_count = 0
        
        for report_dir in self.base_path.iterdir():
            if report_dir.is_dir():
                # Check if directory is old enough
                dir_mtime = report_dir.stat().st_mtime
                if dir_mtime < cutoff_date:
                    try:
                        shutil.rmtree(report_dir)
                        deleted_count += 1
                        logger.info(f"Cleaned up old report: {report_dir.name}")
                    except Exception as e:
                        logger.error(f"Failed to cleanup report {report_dir.name}: {e}")
        
        return deleted_count

class ExcelExporter:
    """Exports validation results to Excel format"""
    
    def __init__(self):
        # Define styles
        self.header_style = {
            'font': Font(bold=True, color='FFFFFF'),
            'fill': PatternFill(start_color='366092', end_color='366092', fill_type='solid'),
            'alignment': Alignment(horizontal='center', vertical='center')
        }
        
        self.mismatch_style = {
            'fill': PatternFill(start_color='FFE6E6', end_color='FFE6E6', fill_type='solid')
        }
        
        self.match_style = {
            'fill': PatternFill(start_color='E6F3E6', end_color='E6F3E6', fill_type='solid')
        }
        
        self.critical_style = {
            'fill': PatternFill(start_color='FF9999', end_color='FF9999', fill_type='solid'),
            'font': Font(bold=True, color='990000')
        }
        
        self.border_style = Border(
            left=Side(style='thin'),
            right=Side(style='thin'),
            top=Side(style='thin'),
            bottom=Side(style='thin')
        )
    
    def export_validation_results(
        self,
        report_name: str,
        summary: ValidationSummary,
        section_summaries: List[SectionSummary],
        comparison_results: List[ComparisonResult],
        calculation_validations: List[CalculationValidation],
        export_path: str
    ) -> str:
        """Export complete validation results to Excel"""
        
        workbook = Workbook()
        
        # Remove default sheet
        workbook.remove(workbook.active)
        
        # Create sheets
        self._create_summary_sheet(workbook, report_name, summary, section_summaries)
        self._create_comparison_results_sheet(workbook, comparison_results)
        self._create_calculation_validation_sheet(workbook, calculation_validations)
        self._create_mismatch_only_sheet(workbook, comparison_results)
        
        # Save workbook
        workbook.save(export_path)
        logger.info(f"Exported validation results to: {export_path}")
        
        return export_path
    
    def _create_summary_sheet(self, workbook: Workbook, report_name: str, summary: ValidationSummary, section_summaries: List[SectionSummary]):
        """Create summary sheet"""
        ws = workbook.create_sheet("Summary", 0)
        
        # Title
        ws['A1'] = f"Validation Report: {report_name}"
        ws['A1'].font = Font(bold=True, size=16)
        ws.merge_cells('A1:D1')
        
        # Timestamp
        ws['A2'] = f"Generated: {datetime.now().strftime('%Y-%m-%d %H:%M:%S')}"
        ws['A2'].font = Font(italic=True)
        ws.merge_cells('A2:D2')
        
        # Overall Summary
        row = 4
        ws[f'A{row}'] = "Overall Summary"
        ws[f'A{row}'].font = self.header_style['font']
        
        row += 1
        summary_data = [
            ("Total Records Compared", summary.total_records_compared),
            ("Total Fields Compared", summary.total_fields_compared),
            ("Total Matches", summary.total_matches),
            ("Total Mismatches", summary.total_mismatches),
            ("Missing in Source", summary.total_missing_in_source),
            ("Missing in Destination", summary.total_missing_in_dest),
            ("Overall Match Percentage", f"{summary.overall_match_percentage}%"),
            ("Critical Issues", summary.critical_issues),
            ("High Issues", summary.high_issues),
            ("Medium Issues", summary.medium_issues),
            ("Low Issues", summary.low_issues),
            ("Has Calculation Errors", "Yes" if summary.has_calculation_errors else "No"),
            ("Has Structural Errors", "Yes" if summary.has_structural_errors else "No")
        ]
        
        for label, value in summary_data:
            ws[f'A{row}'] = label
            ws[f'B{row}'] = value
            
            # Color code critical items
            if "Critical" in label and value > 0:
                ws[f'B{row}'].fill = self.critical_style['fill']
            elif "Match Percentage" in label:
                if float(str(value).replace('%', '')) < 95:
                    ws[f'B{row}'].fill = self.mismatch_style['fill']
                else:
                    ws[f'B{row}'].fill = self.match_style['fill']
            
            row += 1
        
        # Section Summaries
        row += 2
        ws[f'A{row}'] = "Section Summary"
        ws[f'A{row}'].font = self.header_style['font']
        
        row += 1
        section_headers = ["Section", "Records", "Matches", "Mismatches", "Missing (Source)", "Missing (Dest)", "Match %"]
        for col, header in enumerate(section_headers, 1):
            cell = ws.cell(row=row, column=col, value=header)
            cell.font = self.header_style['font']
            cell.fill = self.header_style['fill']
            cell.alignment = self.header_style['alignment']
            cell.border = self.border_style
        
        row += 1
        for section in section_summaries:
            section_data = [
                section.section_name,
                section.total_records,
                section.matches,
                section.mismatches,
                section.missing_in_source,
                section.missing_in_dest,
                f"{section.match_percentage}%"
            ]
            
            for col, value in enumerate(section_data, 1):
                cell = ws.cell(row=row, column=col, value=value)
                cell.border = self.border_style
                
                # Color code match percentage
                if col == 7:  # Match percentage column
                    match_pct = float(str(value).replace('%', ''))
                    if match_pct < 95:
                        cell.fill = self.mismatch_style['fill']
                    else:
                        cell.fill = self.match_style['fill']
            
            row += 1
        
        # Auto-adjust column widths
        for column in ws.columns:
            max_length = 0
            column = [cell for cell in column]
            for cell in column:
                try:
                    if len(str(cell.value)) > max_length:
                        max_length = len(str(cell.value))
                except:
                    pass
            adjusted_width = (max_length + 2)
            ws.column_dimensions[column[0].column_letter].width = adjusted_width
    
    def _create_comparison_results_sheet(self, workbook: Workbook, comparison_results: List[ComparisonResult]):
        """Create detailed comparison results sheet"""
        ws = workbook.create_sheet("Detailed Results")
        
        # Headers
        headers = ["Key", "Section", "Field", "Source Value", "Dest Value", "Status", "Severity", "Difference", "Notes"]
        for col, header in enumerate(headers, 1):
            cell = ws.cell(row=1, column=col, value=header)
            cell.font = self.header_style['font']
            cell.fill = self.header_style['fill']
            cell.alignment = self.header_style['alignment']
            cell.border = self.border_style
        
        # Data
        for row, result in enumerate(comparison_results, 2):
            data = [
                result.key,
                result.section,
                result.field,
                result.source_value,
                result.dest_value,
                result.status.value,
                result.severity.value,
                result.difference,
                result.notes
            ]
            
            for col, value in enumerate(data, 1):
                cell = ws.cell(row=row, column=col, value=value)
                cell.border = self.border_style
                
                # Color code by status
                if result.status.value == "MISMATCH":
                    cell.fill = self.mismatch_style['fill']
                elif result.status.value in ["MISSING_IN_SOURCE", "MISSING_IN_DEST"]:
                    cell.fill = self.critical_style['fill']
                elif result.status.value == "MATCH":
                    cell.fill = self.match_style['fill']
                
                # Bold critical items
                if result.severity.value == "CRITICAL":
                    cell.font = Font(bold=True)
        
        # Auto-adjust column widths
        for column in ws.columns:
            max_length = 0
            column = [cell for cell in column]
            for cell in column:
                try:
                    if len(str(cell.value)) > max_length:
                        max_length = len(str(cell.value))
                except:
                    pass
            adjusted_width = min((max_length + 2), 50)  # Cap at 50 characters
            ws.column_dimensions[column[0].column_letter].width = adjusted_width
    
    def _create_calculation_validation_sheet(self, workbook: Workbook, calculation_validations: List[CalculationValidation]):
        """Create calculation validation sheet"""
        ws = workbook.create_sheet("Calculations")
        
        # Headers
        headers = ["Field", "Expected Value", "Actual Value", "Difference", "Percentage Error", "Status", "Formula Used"]
        for col, header in enumerate(headers, 1):
            cell = ws.cell(row=1, column=col, value=header)
            cell.font = self.header_style['font']
            cell.fill = self.header_style['fill']
            cell.alignment = self.header_style['alignment']
            cell.border = self.border_style
        
        # Data
        for row, calc in enumerate(calculation_validations, 2):
            data = [
                calc.field,
                calc.expected_value,
                calc.actual_value,
                calc.difference,
                f"{calc.percentage_error:.2f}%",
                calc.status.value,
                calc.formula_used
            ]
            
            for col, value in enumerate(data, 1):
                cell = ws.cell(row=row, column=col, value=value)
                cell.border = self.border_style
                
                # Color code by status
                if calc.status.value == "CALCULATION_ERROR":
                    cell.fill = self.critical_style['fill']
                    cell.font = Font(bold=True)
                else:
                    cell.fill = self.match_style['fill']
        
        # Auto-adjust column widths
        for column in ws.columns:
            max_length = 0
            column = [cell for cell in column]
            for cell in column:
                try:
                    if len(str(cell.value)) > max_length:
                        max_length = len(str(cell.value))
                except:
                    pass
            adjusted_width = min((max_length + 2), 40)
            ws.column_dimensions[column[0].column_letter].width = adjusted_width
    
    def _create_mismatch_only_sheet(self, workbook: Workbook, comparison_results: List[ComparisonResult]):
        """Create sheet with only mismatches for quick review"""
        ws = workbook.create_sheet("Mismatches Only")
        
        # Filter mismatches
        mismatches = [r for r in comparison_results if r.status.value != "MATCH"]
        
        if not mismatches:
            ws['A1'] = "No mismatches found - All validations passed!"
            ws['A1'].font = Font(bold=True, color='008000')
            return
        
        # Headers
        headers = ["Key", "Section", "Field", "Source Value", "Dest Value", "Status", "Severity", "Difference", "Notes"]
        for col, header in enumerate(headers, 1):
            cell = ws.cell(row=1, column=col, value=header)
            cell.font = self.header_style['font']
            cell.fill = self.header_style['fill']
            cell.alignment = self.header_style['alignment']
            cell.border = self.border_style
        
        # Data
        for row, result in enumerate(mismatches, 2):
            data = [
                result.key,
                result.section,
                result.field,
                result.source_value,
                result.dest_value,
                result.status.value,
                result.severity.value,
                result.difference,
                result.notes
            ]
            
            for col, value in enumerate(data, 1):
                cell = ws.cell(row=row, column=col, value=value)
                cell.border = self.border_style
                
                # Color code by severity
                if result.severity.value == "CRITICAL":
                    cell.fill = self.critical_style['fill']
                    cell.font = Font(bold=True)
                elif result.severity.value == "HIGH":
                    cell.fill = PatternFill(start_color='FFCC99', end_color='FFCC99', fill_type='solid')
                else:
                    cell.fill = self.mismatch_style['fill']
        
        # Auto-adjust column widths
        for column in ws.columns:
            max_length = 0
            column = [cell for cell in column]
            for cell in column:
                try:
                    if len(str(cell.value)) > max_length:
                        max_length = len(str(cell.value))
                except:
                    pass
            adjusted_width = min((max_length + 2), 50)
            ws.column_dimensions[column[0].column_letter].width = adjusted_width

# Utility functions
def validate_file_extension(filename: str) -> bool:
    """Validate that file has Excel extension"""
    allowed_extensions = ['.xlsx', '.xls']
    return any(filename.lower().endswith(ext) for ext in allowed_extensions)

def get_safe_filename(filename: str) -> str:
    """Get safe filename by removing invalid characters"""
    invalid_chars = ['<', '>', ':', '"', '/', '\\', '|', '?', '*']
    safe_name = filename
    for char in invalid_chars:
        safe_name = safe_name.replace(char, '_')
    return safe_name

def calculate_file_hash(file_path: str) -> str:
    """Calculate MD5 hash of file for integrity checking"""
    import hashlib
    
    hash_md5 = hashlib.md5()
    try:
        with open(file_path, "rb") as f:
            for chunk in iter(lambda: f.read(4096), b""):
                hash_md5.update(chunk)
        return hash_md5.hexdigest()
    except Exception as e:
        logger.error(f"Failed to calculate hash for {file_path}: {e}")
        return ""

def get_directory_size(directory_path: str) -> float:
    """Get total size of directory in MB"""
    total_size = 0
    try:
        for dirpath, dirnames, filenames in os.walk(directory_path):
            for filename in filenames:
                file_path = os.path.join(dirpath, filename)
                if os.path.exists(file_path):
                    total_size += os.path.getsize(file_path)
        return round(total_size / (1024 * 1024), 2)
    except Exception as e:
        logger.error(f"Failed to calculate directory size for {directory_path}: {e}")
        return 0.0